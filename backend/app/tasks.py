"""Celery background tasks for async data processing."""

import logging
import os
import sys
import uuid

# Ensure the root of the project is in python path so we can import etl module
project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

from app.celery_app import celery
from app.database import SessionLocal
from app.models.dataset import Dataset
from app.models.model import ModelVersion
from etl.pipeline import run_pipeline
import joblib
from pathlib import Path
import pandas as pd
from ml.forecasting import prepare_time_series, train_and_evaluate

logger = logging.getLogger(__name__)


@celery.task(name="app.tasks.process_dataset_task")
def process_dataset_task(dataset_id: str) -> str:
    """Asynchronously process an uploaded dataset.

    - Infer schema and validate via Pandera
    - Profile columns (missing, unique, outliers)
    - Check for schema drift compared to previous uploads of same name
    - Clean columns and values, save cleaned copy
    """
    logger.info("Starting background processing for dataset ID: %s", dataset_id)
    db = SessionLocal()
    try:
        db_id = uuid.UUID(dataset_id) if isinstance(dataset_id, str) else dataset_id
        dataset = db.query(Dataset).filter(Dataset.id == db_id).first()
        if not dataset:
            logger.error("Dataset not found: %s", dataset_id)
            return f"Error: Dataset {dataset_id} not found"

        dataset.status = "processing"
        db.commit()

        # Check for previous schema of same file name to detect schema drift
        previous_completed = (
            db.query(Dataset)
            .filter(
                Dataset.user_id == dataset.user_id,
                Dataset.original_filename == dataset.original_filename,
                Dataset.status == "ready",
                Dataset.id != dataset.id,
            )
            .order_by(Dataset.created_at.desc())
            .first()
        )

        previous_schema = None
        if previous_completed and previous_completed.profile_report:
            prev_profile = previous_completed.profile_report
            # Convert dictionary mapping of columns to the list format expected
            if "columns" in prev_profile:
                previous_schema = [
                    {"name": col_name, "dtype": col_data.get("type", "text")}
                    for col_name, col_data in prev_profile["columns"].items()
                ]

        # Execute ETL pipeline
        logger.info("Executing ETL pipeline on file: %s", dataset.file_path)
        report = run_pipeline(dataset.file_path, previous_schema=previous_schema)

        # Update DB record with results
        dataset.status = "ready"
        dataset.profile_report = report
        dataset.cleaned_file_path = report.get("cleaned_file_path")
        dataset.row_count = report["summary"]["total_rows"]
        dataset.column_count = report["summary"]["total_columns"]
        
        # Format columns metadata for backward compatibility with phase 1 view
        cols_meta = []
        for col_name, col_data in report.get("columns", {}).items():
            cols_meta.append({"name": col_name, "dtype": col_data.get("type", "text")})
        dataset.columns_metadata = cols_meta

        db.commit()
        logger.info("Successfully processed dataset: %s", dataset.original_filename)
        return f"Success: Processed dataset {dataset_id}"

    except Exception as exc:
        db.rollback()
        logger.exception("Failed to process dataset %s", dataset_id)
        # Update dataset status to failed
        try:
            dataset = db.query(Dataset).filter(Dataset.id == dataset_id).first()
            if dataset:
                dataset.status = "failed"
                db.commit()
        except Exception:
            logger.exception("Failed to mark dataset as failed in DB")
        return f"Failure: {exc}"
    finally:
        db.close()


@celery.task(name="app.tasks.train_forecast_model_task")
def train_forecast_model_task(dataset_id: str) -> str:
    """Asynchronously train and evaluate a sales forecasting model.

    - Loads clean CSV data using DuckDB/Pandas
    - Aggregates daily sales
    - Engineers time lags and rolling averages
    - Trains XGBoost and evaluates (MAE, RMSE, MAPE)
    - Saves trained model to storage/models/
    - Inserts a ModelVersion row in DB
    """
    logger.info("Starting background training for dataset ID: %s", dataset_id)
    db = SessionLocal()
    try:
        db_id = uuid.UUID(dataset_id) if isinstance(dataset_id, str) else dataset_id
        dataset = db.query(Dataset).filter(Dataset.id == db_id).first()
        if not dataset:
            logger.error("Dataset not found: %s", dataset_id)
            return f"Error: Dataset {dataset_id} not found"

        if not dataset.column_mapping:
            logger.error("Dataset has no column mapping: %s", dataset_id)
            return f"Error: Mapping required for dataset {dataset_id}"

        # 1. Read clean dataset
        # In this phase, we read the cleaned dataset. If not cleaned yet, fallback to raw.
        file_path = dataset.cleaned_file_path if dataset.cleaned_file_path else dataset.file_path
        if not os.path.exists(file_path):
            # Try raw file path
            file_path = dataset.file_path
            if not os.path.exists(file_path):
                return f"Error: CSV file not found on disk at {file_path}"

        df = pd.read_csv(file_path)

        # 2. Extract mapped columns
        mapping = dataset.column_mapping
        date_col = mapping["date_col"]
        # Standardize column naming if cleaned (our auto-cleaning standardizes names to snake_case)
        # Standardize column mapping keys to match cleaned columns if we standardized them
        # Let's search mapping values in df columns. If not present, try standardized key name.
        if date_col not in df.columns:
            # Try standardized name
            clean_date_col = date_col.strip().lower().replace(" ", "_")
            if clean_date_col in df.columns:
                date_col = clean_date_col

        rev_col = mapping["revenue_col"]
        if rev_col not in df.columns:
            clean_rev_col = rev_col.strip().lower().replace(" ", "_")
            if clean_rev_col in df.columns:
                rev_col = clean_rev_col

        df_daily = prepare_time_series(df, date_col, rev_col)

        if len(df_daily) < 10:
            return "Error: Time series must have at least 10 daily points to train a forecasting model."

        # 3. Train & Evaluate XGBoost
        model, metrics, std_error = train_and_evaluate(df_daily, date_col, rev_col)

        # 4. Save model artifact
        models_dir = Path(project_root) / "backend" / "storage" / "models"
        models_dir.mkdir(parents=True, exist_ok=True)

        version_uuid = uuid.uuid4()
        version_id = f"v_{version_uuid.hex[:8]}"
        model_filename = f"{dataset_id}_{version_id}.joblib"
        model_path = models_dir / model_filename

        # Save both model and std_error
        artifact = {"model": model, "std_error": std_error}
        joblib.dump(artifact, model_path)
        logger.info("Saved model artifact to %s", model_path)

        # 5. Insert model run into DB
        model_version = ModelVersion(
            id=version_uuid,
            dataset_id=dataset.id,
            version_id=version_id,
            file_path=str(model_path),
            metrics=metrics,
        )
        db.add(model_version)
        db.commit()

        logger.info("Successfully trained model version: %s", version_id)
        return f"Success: Trained model {version_id} with metrics {metrics}"

    except Exception as exc:
        db.rollback()
        logger.exception("Failed to train forecast model for dataset %s", dataset_id)
        return f"Failure: {exc}"
    finally:
        db.close()


@celery.task(name="app.tasks.generate_pdf_report_task")
def generate_pdf_report_task(
    task_id: str,
    dataset_id: str,
    start_date: str | None,
    end_date: str | None,
    granularity: str,
    category: str | None,
    region: str | None,
) -> str:
    """Asynchronously generate a sales PDF report with structured KPIs, trend charts, and summaries."""
    logger.info("Starting PDF report generation task: %s", task_id)
    db = SessionLocal()
    try:
        import uuid
        from datetime import datetime
        db_id = uuid.UUID(dataset_id) if isinstance(dataset_id, str) else dataset_id
        dataset = db.query(Dataset).filter(Dataset.id == db_id).first()
        if not dataset:
            return f"Error: Dataset {dataset_id} not found"

        # 1. Fetch Analytics data
        from app.services.queries import get_sales_analytics
        file_path = dataset.cleaned_file_path if dataset.cleaned_file_path else dataset.file_path
        analytics = get_sales_analytics(
            file_path=file_path,
            mapping=dataset.column_mapping,
            start_date=start_date,
            end_date=end_date,
            granularity=granularity,
            category_filter=category,
            region_filter=region,
        )

        # 2. Get AI Summary & Recommendations
        from app.services.assistant import summarize_dashboard_service
        kpis = analytics["kpis"]
        summary = summarize_dashboard_service(
            kpis=kpis,
            trend=analytics["trend"],
            categories=analytics["categories"],
            regions=analytics["regions"],
        )
        if hasattr(summary, "__await__"):
            # If running in async environment loop, but Celery task is synchronous:
            import asyncio
            summary = asyncio.run(summary)

        # Generate a list of recommendations based on metrics
        recs = [
            f"Focus marketing efforts on top performing region '{analytics['regions'][0]['region']}' to capitalize on demand.",
            f"Review inventory and stocking levels for leading Category '{analytics['categories'][0]['category']}' to prevent out-of-stock events.",
            "Analyze low revenue periods in the trend chart to plan targeted promotions."
        ]

        # 3. Headless Chart Plotting via Matplotlib
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt

        trend_dates = [t["date"] for t in analytics["trend"]]
        trend_revs = [t["revenue"] for t in analytics["trend"]]

        plt.figure(figsize=(7, 3))
        plt.plot(trend_dates, trend_revs, color="#6366f1", linewidth=2)
        plt.title("Revenue Trend over Time", fontsize=10, fontweight="bold", color="#1e293b")
        plt.xlabel("Date", fontsize=8, color="#64748b")
        plt.ylabel("Revenue ($)", fontsize=8, color="#64748b")
        plt.grid(axis="y", linestyle="--", alpha=0.5)
        plt.xticks(fontsize=6, rotation=30)
        plt.yticks(fontsize=6)
        plt.tight_layout()

        reports_dir = Path(project_root) / "backend" / "storage" / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        chart_path = reports_dir / f"chart_{task_id}.png"
        plt.savefig(chart_path, dpi=200)
        plt.close()

        # 4. Compile PDF Document via ReportLab
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, KeepTogether
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        pdf_path = reports_dir / f"{task_id}.pdf"
        doc = SimpleDocTemplate(
            str(pdf_path),
            pagesize=letter,
            rightMargin=36,
            leftMargin=36,
            topMargin=36,
            bottomMargin=36,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "ReportTitle",
            parent=styles["Heading1"],
            fontSize=22,
            leading=26,
            textColor=colors.HexColor("#6366f1"),
            spaceAfter=15,
        )
        section_style = ParagraphStyle(
            "SectionHeader",
            parent=styles["Heading2"],
            fontSize=14,
            leading=18,
            textColor=colors.HexColor("#1e293b"),
            spaceBefore=12,
            spaceAfter=6,
        )
        body_style = ParagraphStyle(
            "ReportBody",
            parent=styles["Normal"],
            fontSize=10,
            leading=14,
            textColor=colors.HexColor("#475569"),
            spaceAfter=8,
        )

        elements = []

        # Branded Header
        elements.append(Paragraph("Enterprise Decision Intelligence Platform", ParagraphStyle("Sub", fontSize=8, textColor=colors.HexColor("#94a3b8"))))
        elements.append(Paragraph(f"Sales Performance Report — {dataset.original_filename}", title_style))
        elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}", body_style))
        elements.append(Spacer(1, 10))

        # KPI Summary Table
        elements.append(Paragraph("KPI Dashboard Summary", section_style))
        kpi_data = [
            ["Metric", "Value", "Previous Value", "Change %"],
            ["Total Revenue", f"${kpis['revenue']['value']:,.2f}", f"${kpis['revenue']['previous_value']:,.2f}", f"{kpis['revenue']['percentage_change']}%"],
            ["Units Sold", f"{kpis['quantity']['value']:,}", f"{kpis['quantity']['previous_value']:,}", f"{kpis['quantity']['percentage_change']}%"],
            ["Avg Order Value", f"${kpis['aov']['value']:,.2f}", f"${kpis['aov']['previous_value']:,.2f}", f"{kpis['aov']['percentage_change']}%"],
            ["Customer Base", f"{kpis['customers']['value']:,}", f"{kpis['customers']['previous_value']:,}", f"{kpis['customers']['percentage_change']}%"],
        ]
        
        t = Table(kpi_data, colWidths=[150, 120, 120, 100])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
            ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#f8fafc")),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#cbd5e1")),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 15))

        # Trend Chart Image
        elements.append(Paragraph("Historical Revenue Analysis", section_style))
        if os.path.exists(chart_path):
            img = Image(str(chart_path), width=500, height=214)
            elements.append(img)
        elements.append(Spacer(1, 15))

        # Insights
        elements.append(Paragraph("AI Generated Performance Insights", section_style))
        elements.append(Paragraph(summary, body_style))
        elements.append(Spacer(1, 10))

        # Action Recommendations
        elements.append(Paragraph("Recommended Strategic Actions", section_style))
        for rec in recs:
            elements.append(Paragraph(f"• {rec}", body_style))

        # Build PDF
        doc.build(elements)

        # Cleanup chart image
        if os.path.exists(chart_path):
            os.remove(chart_path)

        logger.info("Successfully completed PDF report generation task: %s", task_id)
        return f"Success: PDF generated at {pdf_path}"

    except Exception as exc:
        logger.exception("Failed to build PDF report: %s", task_id)
        return f"Failure: {exc}"
    finally:
        db.close()


@celery.task(name="app.tasks.generate_excel_report_task")
def generate_excel_report_task(
    task_id: str,
    dataset_id: str,
    start_date: str | None,
    end_date: str | None,
    granularity: str,
    category: str | None,
    region: str | None,
) -> str:
    """Asynchronously generate a sales Excel spreadsheet containing formatted KPI metrics and trend rows."""
    logger.info("Starting Excel report generation task: %s", task_id)
    db = SessionLocal()
    try:
        import uuid
        db_id = uuid.UUID(dataset_id) if isinstance(dataset_id, str) else dataset_id
        dataset = db.query(Dataset).filter(Dataset.id == db_id).first()
        if not dataset:
            return f"Error: Dataset {dataset_id} not found"

        # 1. Fetch Analytics data
        from app.services.queries import get_sales_analytics
        file_path = dataset.cleaned_file_path if dataset.cleaned_file_path else dataset.file_path
        analytics = get_sales_analytics(
            file_path=file_path,
            mapping=dataset.column_mapping,
            start_date=start_date,
            end_date=end_date,
            granularity=granularity,
            category_filter=category,
            region_filter=region,
        )

        # 2. Build openpyxl workbook
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()

        # Sheet 1: KPIs
        ws1 = wb.active
        ws1.title = "Executive Summary"
        ws1.views.sheetView[0].showGridLines = True

        # Styles
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=14, bold=True, color="1E293B")
        normal_font = Font(name="Calibri", size=11, color="1E293B")
        bold_font = Font(name="Calibri", size=11, bold=True, color="1E293B")
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")

        thin_side = Side(border_style="thin", color="CBD5E1")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

        align_center = Alignment(horizontal="center", vertical="center")
        align_left = Alignment(horizontal="left", vertical="center")

        ws1["A1"] = "Executive Sales Performance KPI Summary"
        ws1["A1"].font = title_font
        ws1.row_dimensions[1].height = 25

        kpi_headers = ["Metric Descriptor", "Current Period Value", "Previous Period Value", "Change Rate (%)"]
        ws1.append([]) # empty row
        ws1.append(kpi_headers)
        ws1.row_dimensions[3].height = 20

        # Style headers
        for col_idx in range(1, 5):
            cell = ws1.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        kpis = analytics["kpis"]
        kpi_rows = [
            ["Total Revenue", kpis["revenue"]["value"], kpis["revenue"]["previous_value"], kpis["revenue"]["percentage_change"] / 100],
            ["Volume Units Sold", kpis["quantity"]["value"], kpis["quantity"]["previous_value"], kpis["quantity"]["percentage_change"] / 100],
            ["Average Order Value", kpis["aov"]["value"], kpis["aov"]["previous_value"], kpis["aov"]["percentage_change"] / 100],
            ["Total Active Customers", kpis["customers"]["value"], kpis["customers"]["previous_value"], kpis["customers"]["percentage_change"] / 100],
        ]

        for r_idx, row in enumerate(kpi_rows, start=4):
            ws1.append(row)
            ws1.row_dimensions[r_idx].height = 18
            for c_idx in range(1, 5):
                cell = ws1.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                
                # Apply number formatting
                if c_idx == 1:
                    cell.alignment = align_left
                elif c_idx in (2, 3):
                    cell.alignment = align_center
                    if r_idx in (4, 6): # revenue and AOV are currencies
                        cell.number_format = "$#,##0.00"
                    else: # quantity and customers are counts
                        cell.number_format = "#,##0"
                elif c_idx == 4:
                    cell.alignment = align_center
                    cell.number_format = "0.0%"

        # Auto-adjust column widths
        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # Sheet 2: Raw Trend Aggregates
        ws2 = wb.create_sheet(title="Daily Sales Trend")
        ws2.views.sheetView[0].showGridLines = True

        ws2["A1"] = f"Granular Daily Sales Trend ({granularity})"
        ws2["A1"].font = title_font
        ws2.row_dimensions[1].height = 25
        ws2.append([])

        trend_headers = ["Sequence Date", "Revenue Turnover ($)", "Units Sold Volume"]
        ws2.append(trend_headers)
        ws2.row_dimensions[3].height = 20

        for col_idx in range(1, 4):
            cell = ws2.cell(row=3, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_center

        for r_idx, t_point in enumerate(analytics["trend"], start=4):
            ws2.append([t_point["date"], t_point["revenue"], t_point["quantity"]])
            ws2.row_dimensions[r_idx].height = 16
            for c_idx in range(1, 4):
                cell = ws2.cell(row=r_idx, column=c_idx)
                cell.font = normal_font
                cell.border = border
                if r_idx % 2 == 1:
                    cell.fill = zebra_fill
                
                if c_idx == 1:
                    cell.alignment = align_center
                elif c_idx == 2:
                    cell.alignment = align_center
                    cell.number_format = "$#,##0.00"
                elif c_idx == 3:
                    cell.alignment = align_center
                    cell.number_format = "#,##0"

        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 4, 15)

        # Save workbook to disk
        reports_dir = Path(project_root) / "backend" / "storage" / "reports"
        reports_dir.mkdir(parents=True, exist_ok=True)
        excel_path = reports_dir / f"{task_id}.xlsx"
        wb.save(excel_path)

        logger.info("Successfully completed Excel report generation task: %s", task_id)
        return f"Success: Excel workbook generated at {excel_path}"

    except Exception as exc:
        logger.exception("Failed to build Excel report: %s", task_id)
        return f"Failure: {exc}"
    finally:
        db.close()


@celery.task(name="app.tasks.train_segmentation_model_task")
def train_segmentation_model_task(dataset_id: str) -> str:
    """Asynchronously pre-process RFM features and train a customer K-Means segmentation model."""
    logger.info("Starting background K-Means training for dataset ID: %s", dataset_id)
    db = SessionLocal()
    try:
        import uuid
        db_id = uuid.UUID(dataset_id) if isinstance(dataset_id, str) else dataset_id
        dataset = db.query(Dataset).filter(Dataset.id == db_id).first()
        if not dataset:
            logger.error("Dataset not found: %s", dataset_id)
            return f"Error: Dataset {dataset_id} not found"

        file_path = dataset.cleaned_file_path if dataset.cleaned_file_path else dataset.file_path
        if not file_path or not os.path.exists(file_path):
            return f"Error: File path {file_path} not found"

        df = pd.read_csv(file_path)

        # 1. Feature Preprocessing
        from ml.segmentation import prepare_segmentation_features, train_segmentation
        df_features = prepare_segmentation_features(df, dataset.column_mapping)

        # 2. Model Training
        kmeans, scaler, metrics, df_labeled = train_segmentation(df_features)

        # 3. Save Model Run Artifacts
        version_uuid = uuid.uuid4()
        version_id = f"seg_v_{version_uuid.hex[:8]}"
        
        models_dir = Path(project_root) / "backend" / "storage" / "models"
        models_dir.mkdir(parents=True, exist_ok=True)
        model_path = models_dir / f"{version_id}.joblib"

        joblib.dump({"model": kmeans, "scaler": scaler, "metrics": metrics}, model_path)
        logger.info("Saved model artifact to %s", model_path)

        # 4. Insert model run record into DB
        # Convert metrics dict to floats/ints for DB storage compatibility
        model_metrics = {
            "silhouette": metrics["silhouette"],
            "cluster_sizes": {str(k): int(v) for k, v in metrics["cluster_sizes"].items()},
            # Keep profiles short/safe for db metadata columns
            "profiles": metrics["profiles"]
        }

        model_version = ModelVersion(
            id=version_uuid,
            dataset_id=dataset.id,
            version_id=version_id,
            file_path=str(model_path),
            metrics=model_metrics,
        )
        db.add(model_version)
        db.commit()

        # Save labeled customer assignments directly back to the database as JSON or a temporary schema
        # For simplicity, we can serialize the first 100 customer labels to model metadata
        sample_customers = []
        for _, row in df_labeled.head(100).iterrows():
            sample_customers.append({
                "customer": str(row.iloc[0]), # index 0 is customer ID
                "spend": float(row["total_spend"]),
                "frequency": int(row["frequency"]),
                "segment": row["segment_name"]
            })
        
        # Save to metadata columns
        model_version.metrics["sample_customers"] = sample_customers
        db.add(model_version)
        db.commit()

        logger.info("Successfully trained customer segmentation version: %s", version_id)
        return f"Success: Trained K-Means model {version_id} with silhouette {metrics['silhouette']}%"

    except Exception as exc:
        db.rollback()
        logger.exception("Failed to train customer segments for dataset %s", dataset_id)
        return f"Failure: {exc}"
    finally:
        db.close()



