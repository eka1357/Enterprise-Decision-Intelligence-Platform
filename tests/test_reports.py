"""Tests for PDF, Excel, and CSV Report Export routers and tasks."""

import os
import uuid
import pytest
from fastapi.testclient import TestClient
from sqlalchemy.orm import Session
import openpyxl

from app.models.dataset import Dataset
from app.tasks import generate_pdf_report_task, generate_excel_report_task


@pytest.fixture()
def sales_csv(tmp_path) -> str:
    """Create a mock sales CSV dataset."""
    csv_content = (
        "date,revenue,quantity,category,region\n"
        "2026-07-01,150.0,3,Technology,North\n"
        "2026-07-02,80.0,2,Furniture,South\n"
        "2026-07-03,20.0,1,Office Supplies,East\n"
        "2026-07-04,300.0,5,Technology,West\n"
        "2026-07-05,120.0,4,Office Supplies,North\n"
    )
    csv_file = tmp_path / "sales_sample.csv"
    csv_file.write_text(csv_content)
    return str(csv_file)


@pytest.fixture()
def mock_dataset_for_reports(db: Session, registered_user: dict, sales_csv: str) -> Dataset:
    """Create a mock dataset with valid column mapping for report tests."""
    user_id = registered_user["response"]["id"]
    dataset = Dataset(
        id=uuid.uuid4(),
        user_id=uuid.UUID(user_id),
        filename="sales_sample.csv",
        original_filename="sales_sample.csv",
        file_path=sales_csv,
        file_size_bytes=2048,
        row_count=5,
        column_count=5,
        status="ready",
        column_mapping={
            "date_col": "date",
            "revenue_col": "revenue",
            "quantity_col": "quantity",
            "category_col": "category",
            "region_col": "region",
        },
    )
    db.add(dataset)
    db.commit()
    db.refresh(dataset)
    return dataset


def test_pdf_report_generation_task(mock_dataset_for_reports: Dataset) -> None:
    """Verifies that the PDF generator task compiles a non-empty PDF binary."""
    task_id = f"test_pdf_{uuid.uuid4()}"
    
    # Run task function synchronously
    res = generate_pdf_report_task(
        task_id=task_id,
        dataset_id=str(mock_dataset_for_reports.id),
        start_date=None,
        end_date=None,
        granularity="daily",
        category=None,
        region=None,
    )
    
    assert "Success" in res
    
    # Assert file exists and contains PDF signature %PDF
    pdf_path = os.path.join(os.path.dirname(__file__), "..", "backend", "storage", "reports", f"{task_id}.pdf")
    assert os.path.exists(pdf_path)
    assert os.path.getsize(pdf_path) > 0
    
    with open(pdf_path, "rb") as f:
        header = f.read(4)
        assert header == b"%PDF"
        
    # Cleanup
    os.remove(pdf_path)


def test_excel_report_generation_task(mock_dataset_for_reports: Dataset) -> None:
    """Verifies that the Excel task compiles a valid openpyxl spreadsheet workbook."""
    task_id = f"test_excel_{uuid.uuid4()}"
    
    res = generate_excel_report_task(
        task_id=task_id,
        dataset_id=str(mock_dataset_for_reports.id),
        start_date=None,
        end_date=None,
        granularity="daily",
        category=None,
        region=None,
    )
    
    assert "Success" in res
    
    excel_path = os.path.join(os.path.dirname(__file__), "..", "backend", "storage", "reports", f"{task_id}.xlsx")
    assert os.path.exists(excel_path)
    
    # Load and inspect spreadsheet structure
    wb = openpyxl.load_workbook(excel_path)
    assert "Executive Summary" in wb.sheetnames
    assert "Daily Sales Trend" in wb.sheetnames
    
    ws1 = wb["Executive Summary"]
    assert ws1["A1"].value == "Executive Sales Performance KPI Summary"
    assert ws1["A3"].value == "Metric Descriptor"
    
    ws2 = wb["Daily Sales Trend"]
    assert ws2["A3"].value == "Sequence Date"
    
    wb.close()
    # Cleanup
    os.remove(excel_path)


def test_reports_api_endpoints(
    client: TestClient, auth_headers: dict, mock_dataset_for_reports: Dataset
) -> None:
    """Verifies reports routes accept queries and trigger async jobs or stream streams."""
    payload = {
        "start_date": None,
        "end_date": None,
        "granularity": "weekly",
        "category": "Office Supplies",
        "region": "Central",
    }
    
    # 1. Trigger PDF Export
    res_pdf = client.post(
        f"/api/v1/datasets/{mock_dataset_for_reports.id}/reports/pdf",
        json=payload,
        headers=auth_headers,
    )
    assert res_pdf.status_code == 202
    assert "task_id" in res_pdf.json()
    assert res_pdf.json()["status"] == "processing"
    
    # 2. Trigger Excel Export
    res_xls = client.post(
        f"/api/v1/datasets/{mock_dataset_for_reports.id}/reports/excel",
        json=payload,
        headers=auth_headers,
    )
    assert res_xls.status_code == 202
    assert "task_id" in res_xls.json()
    
    # 3. Stream CSV directly
    res_csv = client.get(
        f"/api/v1/datasets/{mock_dataset_for_reports.id}/reports/csv?granularity=weekly",
        headers=auth_headers,
    )
    assert res_csv.status_code == 200
    assert "text/csv" in res_csv.headers["content-type"]
    assert "Date,Revenue,Units Quantity" in res_csv.text
